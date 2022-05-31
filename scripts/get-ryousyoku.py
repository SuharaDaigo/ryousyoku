%apt install ghostscript

#献立を取得、エクセルに置き換え
import camelot
import pandas as pd
import glob
import openpyxl

from tqdm.notebook import tqdm
from more_itertools import chunked

tables = camelot.read_pdf(
  'https://www.maizuru-ct.ac.jp/wp-content/uploads/sites/3/2022/05/menu_domitory_2022-06.pdf',
  pages = 'all'     # ページによってはメモリ使用量が跳ね上がるので1ページづつ解析した方がよい
)
tables

i=0
while (i<6):

  tables[i] #[]の中がページ数([0]=1ページ)

  ftable=tables[i]
  ftable.df #データフレーム

  ftable.to_excel("ryousyoku{}.xlsx".format(i),index=False)  #エクセルファイルに

  #4つに分割したエクセルファイルを一つにまとめるプログラム

files=glob.glob("ryousyoku*.xlsx")

list=[]

for file in files:
  list.append(pd.read_excel(file))

df = pd.concat(list,axis=1)  

df.to_excel("ryousyoku.xlsx",index=False)

#openpyxlモジュールを使用する
import openpyxl
 
#既存のExcelファイルを開く
wb=openpyxl.load_workbook('ryousyoku.xlsx')
#既存ファイルのシートを指定
sheet1=wb['Sheet1']
#既存ファイルに新規シートをシート名と位置を指定して作成
sheet2=wb.create_sheet(title='Sheet2',index=1)
 
#Sheet1の値のある行数を取得
rw=sheet1.max_row
#Sheet1の値のある列数を取得
cl=sheet1.max_column
 
#iは値のある行数分繰り返す
#jは値のある列数分繰り返す
#range(start,stop)はstart≦i<stopでstopで指定した値は含まないので「+1」している
for i in range(1,rw+1):
    for j in range(1,cl+1):
        C1=sheet1.cell(row=i,column=j) #sheet1のセルの行番号と列番号を指定している
        C2=sheet2.cell(row=j,column=i) #sheet1のセルの行番号と列番号を入れ替えてsheet2のセルを指定している
        C2.value=C1.value #sheet2のセルにsheet1のセルの値を代入
 
wb.save('ryousyoku.xlsx') #上書き保存
